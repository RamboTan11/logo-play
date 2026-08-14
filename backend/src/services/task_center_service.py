"""Administrator-only task assignment, image delivery, and export workflow."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from uuid import uuid4
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import Select, func, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from src.config import AppSettings
from src.db.models import AssetRecord, Customer, DesignTask
from src.models.customer_decision import (
    AdminDesignTaskDetailDto,
    AdminDesignTaskDetailResponseDto,
    AdminDesignTaskListDto,
    AdminDesignTaskListItemDto,
)
from src.services.asset_service import (
    LOCAL_FALLBACK,
    TASK_DELIVERY_IMAGE_PURPOSE,
    LocalFallbackAssetStorage,
)
from src.services.customer_access_service import CustomerAccessService
from src.services.event_service import EventService
from src.services.lark_notification_service import LarkWorkflowService

_ALLOWED_STATUSES = {"waiting_assignment", "in_progress", "completed", "canceled"}
_MAX_DELIVERY_BYTES = 10 * 1024 * 1024
_BEIJING_TZ = ZoneInfo("Asia/Shanghai")
_DELIVERY_MEDIA_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
}


class TaskCenterError(RuntimeError):
    """A stable, safe error response for administrator task operations."""

    def __init__(self, code: str, message: str, status_code: int) -> None:
        super().__init__(message)
        self.code = code
        self.status_code = status_code


@dataclass(frozen=True, slots=True)
class DeliveryImageInput:
    """Validated upload metadata; bytes never enter a response or audit payload."""

    filename: str
    media_type: str
    content: bytes


@dataclass(frozen=True, slots=True)
class TaskExportRow:
    """Task data and private image bytes needed only while creating the workbook."""

    customer_name: str
    domain: str
    submitted_at: datetime
    adoption_suggestion: str | None
    adopted_image: bytes | None
    delivery_image: bytes | None
    delivery_placeholder: str


class TaskCenterService:
    """Apply the three-state task machine and protect delivery assets."""

    def __init__(self, asset_root: str, settings: AppSettings, events: EventService | None = None) -> None:
        self._storage = LocalFallbackAssetStorage(asset_root)
        self._events = events or EventService()
        self._lark = LarkWorkflowService(self._events)
        self._customer_access = CustomerAccessService(settings, events=self._events)

    async def list_tasks(
        self,
        session: AsyncSession,
        *,
        statuses: list[str],
        submitted_from: date | None,
        submitted_to: date | None,
        page: int,
        page_size: int,
    ) -> AdminDesignTaskListDto:
        statement = self._filtered_statement(statuses, submitted_from, submitted_to)
        total = int(
            await session.scalar(select(func.count()).select_from(statement.subquery())) or 0
        )
        rows = (
            await session.execute(
                statement.order_by(DesignTask.submitted_at.desc(), DesignTask.id.desc())
                .offset((page - 1) * page_size)
                .limit(page_size)
            )
        ).all()
        return AdminDesignTaskListDto(
            items=[self._summary(task, customer_name, customer) for task, customer_name, customer in rows],
            total=total,
            page=page,
            page_size=page_size,
        )

    async def detail(
        self, session: AsyncSession, task_id: str
    ) -> AdminDesignTaskDetailResponseDto:
        task, customer_name, customer = await self._task_with_customer_or_error(session, task_id)
        return AdminDesignTaskDetailResponseDto(
            task=AdminDesignTaskDetailDto(
                **self._summary(task, customer_name, customer).model_dump(),
                adopted_image_url=f"/api/v1/design-tasks/{task.id}/adopted-image/content",
                delivery_image_url=(
                    f"/api/v1/design-tasks/{task.id}/delivery-image/content"
                    if task.delivery_asset_id is not None
                    else None
                ),
            )
        )

    async def accept(
        self, session: AsyncSession, *, task_id: str, administrator_id: str
    ) -> AdminDesignTaskListItemDto:
        task = await self._task_or_error(session, task_id)
        await self._require_customer_access(session, task)
        if task.status == "canceled":
            raise TaskCenterError(
                "task_changed", "Task was replaced; refresh the task list", 409
            )
        if task.status == "completed":
            raise TaskCenterError("task_already_completed", "Closed task cannot be accepted", 409)
        if task.status == "in_progress":
            return await self._summary_for_task(session, task)
        now = datetime.now(UTC)
        # Re-read immediately before the state transition so a stale admin
        # page cannot win a customer stop/expiry race.
        await self._require_customer_access(session, task)
        result = await session.execute(
            update(DesignTask)
            .where(
                DesignTask.id == task_id,
                DesignTask.status == "waiting_assignment",
                DesignTask.customer_id.in_(self._active_customer_ids(current=now)),
            )
            .values(status="in_progress", updated_at=now)
        )
        if getattr(result, "rowcount", None) != 1:
            refreshed = await self._task_or_error(session, task_id)
            await self._require_customer_access(session, refreshed)
            if refreshed.status == "in_progress":
                return await self._summary_for_task(session, refreshed)
            raise TaskCenterError("task_changed", "Task state changed; refresh the task list", 409)
        task.status = "in_progress"
        task.updated_at = now
        trace_id = uuid4().hex
        await self._events.record_audit(
            session,
            action="design_task.accepted",
            resource_type="design_task",
            resource_id=task.id,
            actor_id=administrator_id,
            trace_id=trace_id,
            summary={"status": "in_progress"},
        )
        await self._events.enqueue_notification(
            session,
            event_type="task.accepted",
            resource_type="design_task",
            resource_id=task.id,
            trace_id=trace_id,
            payload={"task_id": task.id},
        )
        await self._lark.stop_task(
            session,
            task_id=task.id,
            event_type="task.waiting_assignment_overdue",
            reason="accepted",
            now=now,
        )
        await self._lark.snapshot_stage(
            session,
            task_id=task.id,
            event_type="task.upload_overdue",
            entered_at=now,
        )
        return await self._summary_for_task(session, task)

    async def deliver(
        self,
        session: AsyncSession,
        *,
        task_id: str,
        administrator_id: str,
        upload: DeliveryImageInput,
    ) -> AdminDesignTaskListItemDto:
        task = await self._task_or_error(session, task_id)
        await self._require_customer_access(session, task)
        if task.status != "in_progress":
            if task.status == "canceled":
                raise TaskCenterError(
                    "task_changed", "Task was replaced; refresh the task list", 409
                )
            code = "task_already_completed" if task.status == "completed" else "task_not_in_progress"
            raise TaskCenterError(code, "Task must be in progress before delivery", 409)

        now = datetime.now(UTC)
        stored = self._storage.write(upload.content, upload.media_type)
        # Delivery must remain successful even if the optional preview cannot
        # be generated. Historical assets still use the guarded lazy fallback.
        self._storage.write_thumbnail(stored.storage_key, upload.content)
        asset = AssetRecord(
            asset_id=stored.asset_id,
            purpose=TASK_DELIVERY_IMAGE_PURPOSE,
            storage_backend=LOCAL_FALLBACK,
            storage_key=stored.storage_key,
            content_hash=stored.content_hash,
            media_type=upload.media_type,
            size=len(upload.content),
            original_filename=upload.filename,
            source_resource_type="design_task",
            source_resource_id=task_id,
        )
        session.add(asset)
        try:
            await session.flush()
            # The first check protects the upload path before storage. This
            # second check closes the small state-change window before commit.
            await self._require_customer_access(session, task)
            result = await session.execute(
                update(DesignTask)
                .where(
                    DesignTask.id == task_id,
                    DesignTask.status == "in_progress",
                    DesignTask.delivery_asset_id.is_(None),
                    DesignTask.customer_id.in_(self._active_customer_ids(current=now)),
                )
                .values(
                    status="completed",
                    delivery_asset_id=asset.asset_id,
                    updated_at=now,
                )
            )
            if getattr(result, "rowcount", None) != 1:
                await session.delete(asset)
                await session.flush()
                self._storage.delete(stored.storage_key)
                await self._require_customer_access(session, task)
                raise TaskCenterError(
                    "task_changed", "Task state changed; refresh the task list", 409
                )
        except Exception:
            self._storage.delete(stored.storage_key)
            raise

        task.status = "completed"
        task.delivery_asset_id = asset.asset_id
        task.updated_at = now
        trace_id = uuid4().hex
        await self._events.record_audit(
            session,
            action="design_task.delivered",
            resource_type="design_task",
            resource_id=task.id,
            actor_id=administrator_id,
            trace_id=trace_id,
            summary={
                "status": "completed",
                "delivery_media_type": upload.media_type,
                "delivery_size": len(upload.content),
                "storage_backend": LOCAL_FALLBACK,
            },
        )
        await self._events.enqueue_notification(
            session,
            event_type="task.delivered",
            resource_type="design_task",
            resource_id=task.id,
            trace_id=trace_id,
            payload={"task_id": task.id},
        )
        await self._lark.stop_task(
            session,
            task_id=task.id,
            reason="completed",
            now=now,
        )
        await self._lark.enqueue_immediate(
            session,
            event_type="task.delivery_uploaded",
            task_id=task.id,
            trace_id=trace_id,
        )
        return await self._summary_for_task(session, task)

    async def read_adopted_image(
        self, session: AsyncSession, task_id: str
    ) -> tuple[str, bytes]:
        task = await self._task_or_error(session, task_id)
        return await self._read_asset(session, task.adopted_asset_id)

    async def read_delivery_image(
        self, session: AsyncSession, task_id: str
    ) -> tuple[str, bytes]:
        task = await self._task_or_error(session, task_id)
        return await self._read_asset(session, task.delivery_asset_id, delivery=True)

    async def export_tasks(
        self,
        session: AsyncSession,
        *,
        statuses: list[str],
        submitted_from: date | None,
        submitted_to: date | None,
    ) -> bytes:
        rows = (
            await session.execute(
                self._filtered_statement(statuses, submitted_from, submitted_to).order_by(
                    DesignTask.submitted_at.desc(), DesignTask.id.desc()
                )
            )
        ).all()
        export_rows: list[TaskExportRow] = []
        for task, customer_name, _customer in rows:
            adopted_image = await self._export_image_or_none(session, task.adopted_asset_id)
            delivery_image = await self._export_image_or_none(
                session, task.delivery_asset_id, delivery=True
            )
            export_rows.append(
                TaskExportRow(
                    customer_name=customer_name,
                    domain=task.domain,
                    submitted_at=_as_utc(task.submitted_at),
                    adoption_suggestion=task.adoption_suggestion,
                    adopted_image=adopted_image,
                    delivery_image=delivery_image,
                    delivery_placeholder=(
                        "待上传" if task.delivery_asset_id is None else "图片不可用"
                    ),
                )
            )
        return _xlsx_bytes(export_rows)

    def _filtered_statement(
        self, statuses: list[str], submitted_from: date | None, submitted_to: date | None
    ) -> Select[tuple[DesignTask, str, Customer]]:
        invalid = set(statuses).difference(_ALLOWED_STATUSES)
        if invalid:
            raise TaskCenterError("invalid_task_status", "Invalid task status filter", 422)
        statement = (
            select(DesignTask, Customer.name, Customer)
            .join(Customer, Customer.id == DesignTask.customer_id)
            .where(DesignTask.adopted_asset_id.is_not(None))
        )
        if statuses:
            statement = statement.where(DesignTask.status.in_(statuses))
        if submitted_from is not None:
            statement = statement.where(
                DesignTask.submitted_at
                >= datetime.combine(submitted_from, time.min, _BEIJING_TZ).astimezone(UTC)
            )
        if submitted_to is not None:
            statement = statement.where(
                DesignTask.submitted_at
                < datetime.combine(
                    submitted_to + timedelta(days=1), time.min, _BEIJING_TZ
                ).astimezone(UTC)
            )
        return statement

    @staticmethod
    def _active_customer_ids(*, current: datetime) -> Select[tuple[str]]:
        return select(Customer.id).where(
            Customer.access_state == "active",
            Customer.access_expires_at.is_not(None),
            Customer.access_expires_at > current,
        )

    async def _task_or_error(self, session: AsyncSession, task_id: str) -> DesignTask:
        task = await session.scalar(select(DesignTask).where(DesignTask.id == task_id))
        if task is None or task.adopted_asset_id is None:
            raise TaskCenterError("design_task_not_found", "Task not found", 404)
        return task

    async def _task_with_customer_or_error(
        self, session: AsyncSession, task_id: str
    ) -> tuple[DesignTask, str, Customer]:
        row = (
            await session.execute(
                select(DesignTask, Customer.name, Customer)
                .join(Customer, Customer.id == DesignTask.customer_id)
                .where(DesignTask.id == task_id, DesignTask.adopted_asset_id.is_not(None))
            )
        ).one_or_none()
        if row is None:
            raise TaskCenterError("design_task_not_found", "Task not found", 404)
        task, customer_name, customer = row
        return task, customer_name, customer

    async def _summary_for_task(
        self, session: AsyncSession, task: DesignTask
    ) -> AdminDesignTaskListItemDto:
        row = (await session.execute(select(Customer.name, Customer).where(Customer.id == task.customer_id))).one_or_none()
        if row is None:
            raise TaskCenterError("design_task_not_found", "Task not found", 404)
        customer_name, customer = row
        return self._summary(task, customer_name, customer)

    async def _require_customer_access(self, session: AsyncSession, task: DesignTask) -> None:
        customer = await session.get(Customer, task.customer_id)
        if customer is None:
            raise TaskCenterError("design_task_not_found", "Task not found", 404)
        access_status = self._customer_access.derive_status(customer)
        if access_status == "expired":
            raise TaskCenterError("customer_access_expired", "该用户的访客链接已到期，请先恢复使用", 409)
        if access_status == "stopped":
            raise TaskCenterError("customer_access_stopped", "该用户的访客链接已关停，请先恢复使用", 409)
        if access_status == "unstarted":
            raise TaskCenterError("customer_access_unstarted", "Customer access is not started", 409)

    async def _read_asset(
        self, session: AsyncSession, asset_id: str | None, *, delivery: bool = False
    ) -> tuple[str, bytes]:
        if asset_id is None:
            raise TaskCenterError("task_delivery_not_found", "Task delivery image not found", 404)
        asset = await session.get(AssetRecord, asset_id)
        expected_purpose = TASK_DELIVERY_IMAGE_PURPOSE if delivery else "generated_logo"
        if asset is None or asset.purpose != expected_purpose:
            raise TaskCenterError("task_delivery_not_found", "Task delivery image not found", 404)
        try:
            return asset.media_type, self._storage.read(asset.storage_key)
        except OSError as error:
            raise TaskCenterError("task_delivery_not_found", "Task delivery image not found", 404) from error

    async def _export_image_or_none(
        self, session: AsyncSession, asset_id: str | None, *, delivery: bool = False
    ) -> bytes | None:
        if asset_id is None:
            return None
        try:
            _, content = await self._read_asset(session, asset_id, delivery=delivery)
        except TaskCenterError:
            return None
        return content

    def _summary(self, task: DesignTask, customer_name: str, customer: Customer) -> AdminDesignTaskListItemDto:
        return AdminDesignTaskListItemDto(
            id=task.id,
            customer_name=customer_name,
            domain=task.domain,
            status=task.status,
            adoption_suggestion=task.adoption_suggestion,
            submitted_at=_as_utc(task.submitted_at),
            customer_access_status=self._customer_access.derive_status(customer),
        )


def _as_utc(value: datetime) -> datetime:
    """Restore the UTC timezone SQLite removes before serializing task timestamps."""

    return value.replace(tzinfo=UTC) if value.tzinfo is None else value.astimezone(UTC)


def validate_delivery_upload(filename: str | None, media_type: str | None, content: bytes) -> DeliveryImageInput:
    """Require matching extension, declared MIME, signature, and the V1 size limit."""

    safe_name = Path(filename or "").name
    suffix = Path(safe_name).suffix.lower()
    expected_type = _DELIVERY_MEDIA_TYPES.get(suffix)
    if expected_type is None:
        raise TaskCenterError("delivery_image_extension_invalid", "Only PNG or JPEG images are allowed", 422)
    if not content or len(content) > _MAX_DELIVERY_BYTES:
        raise TaskCenterError("delivery_image_size_invalid", "Image must be no larger than 10 MB", 422)
    declared = (media_type or "").lower().split(";", 1)[0].strip()
    if declared != expected_type:
        raise TaskCenterError("delivery_image_mime_invalid", "Image MIME type does not match filename", 422)
    actual = _signature_media_type(content)
    if actual is None or actual != expected_type:
        raise TaskCenterError("delivery_image_signature_invalid", "Image content is invalid", 422)
    return DeliveryImageInput(filename=safe_name, media_type=actual, content=content)


def _signature_media_type(content: bytes) -> str | None:
    if content.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if len(content) >= 4 and content.startswith(b"\xff\xd8\xff") and content.endswith(b"\xff\xd9"):
        return "image/jpeg"
    return None


def _xlsx_bytes(rows: list[TaskExportRow]) -> bytes:
    """Create a readable task workbook with private image thumbnails embedded in it."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "任务中心"
    worksheet.freeze_panes = "A2"
    worksheet.append(["客户名称", "域名", "提交时间", "人工精修建议", "客户选择图片", "精修终稿"])

    header_fill = PatternFill("solid", fgColor="E8ECEF")
    header_font = Font(bold=True, color="25313B")
    center = Alignment(horizontal="center", vertical="center")
    thin_gray = Side(style="thin", color="D8DEE3")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    placeholder_fill = PatternFill("solid", fgColor="F1F3F5")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row_index, row in enumerate(rows, start=2):
        submitted_at = row.submitted_at.astimezone(_BEIJING_TZ).replace(tzinfo=None)
        worksheet.append(
            [
                row.customer_name,
                row.domain,
                submitted_at,
                row.adoption_suggestion or "-",
                None,
                None,
            ]
        )
        worksheet.row_dimensions[row_index].height = 78
        for cell in worksheet[row_index]:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        worksheet.cell(row_index, 3).number_format = "yyyy-mm-dd hh:mm"
        _add_thumbnail(worksheet, row.adopted_image, row_index, 5, "图片不可用", placeholder_fill, border)
        _add_thumbnail(
            worksheet,
            row.delivery_image,
            row_index,
            6,
            row.delivery_placeholder,
            placeholder_fill,
            border,
        )

    for column, width in {"A": 18, "B": 24, "C": 19, "D": 30, "E": 18, "F": 18}.items():
        worksheet.column_dimensions[column].width = width
    worksheet.row_dimensions[1].height = 24
    worksheet.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _add_thumbnail(
    worksheet: Worksheet,
    content: bytes | None,
    row_index: int,
    column_index: int,
    placeholder: str,
    placeholder_fill: PatternFill,
    border: Border,
) -> None:
    """Keep image cells usable when an asset is absent or its bytes are no longer readable."""

    cell = worksheet.cell(row_index, column_index)
    if content is None:
        cell.value = placeholder
        cell.fill = placeholder_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        return
    try:
        image = ExcelImage(BytesIO(content))
    except OSError:
        cell.value = placeholder
        cell.fill = placeholder_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        return
    image.width, image.height = _thumbnail_size(image.width, image.height)
    image.anchor = cell.coordinate
    worksheet.add_image(image)


def _thumbnail_size(width: int, height: int, *, limit: int = 72) -> tuple[int, int]:
    scale = min(limit / width, limit / height, 1)
    return max(1, round(width * scale)), max(1, round(height * scale))
